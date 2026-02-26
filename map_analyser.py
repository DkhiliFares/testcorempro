"""
IAR EWARM Map File Analyzer — STM32 Benchmarking Tool
=====================================================
A modern GUI application to parse IAR EWARM .map files,
extract total and per-function memory footprint (Flash/RAM),
and export results to Excel for benchmarking STM32 vs competitors.

Dependencies: pip install openpyxl
"""

import re
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ─────────────────────────────────────────────────────────────────────────────
# THEME COLORS
# ─────────────────────────────────────────────────────────────────────────────
COLORS = {
    "bg_dark":        "#0f1117",
    "bg_panel":       "#1a1d27",
    "bg_card":        "#232736",
    "bg_input":       "#2a2e3d",
    "accent":         "#6c63ff",
    "accent_hover":   "#7f78ff",
    "accent_green":   "#2ecc71",
    "accent_orange":  "#f39c12",
    "accent_red":     "#e74c3c",
    "text_primary":   "#e8e8ec",
    "text_secondary": "#9b9bb0",
    "text_muted":     "#6c6c80",
    "border":         "#2f3347",
    "success":        "#27ae60",
    "highlight_row":  "#2a2e4a",
}

FONT_FAMILY = "Segoe UI"


# ─────────────────────────────────────────────────────────────────────────────
# MAP FILE PARSER
# ─────────────────────────────────────────────────────────────────────────────
def parse_iar_number(s: str) -> int:
    """Parse IAR-formatted numbers like 10'751 or 1'168."""
    if not s or s.strip() == "":
        return 0
    return int(s.strip().replace("'", "").replace(",", ""))


def parse_map_file(filepath: str) -> dict:
    """
    Parse an IAR EWARM .map file and return structured data.

    Returns dict with keys:
        - project_name: str
        - toolchain_info: str
        - modules: list of dicts {name, ro_code, ro_data, rw_data}
        - entries: list of dicts {name, address, size, type, object}
        - grand_total: dict {ro_code, ro_data, rw_data}
        - summary: dict {readonly_code, readonly_data, readwrite_data}
    """
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        content = f.read()
    lines = content.splitlines()

    result = {
        "project_name": os.path.splitext(os.path.basename(filepath))[0],
        "toolchain_info": "",
        "modules": [],
        "entries": [],
        "grand_total": {"ro_code": 0, "ro_data": 0, "rw_data": 0},
        "summary": {"readonly_code": 0, "readonly_data": 0, "readwrite_data": 0},
    }

    # ── Extract toolchain info ──────────────────────────────────────────
    for line in lines[:10]:
        if "IAR ELF Linker" in line:
            result["toolchain_info"] = line.strip().lstrip("#").strip()
            break

    # ── Locate sections ─────────────────────────────────────────────────
    module_summary_start = None
    entry_list_start = None

    for i, line in enumerate(lines):
        if "MODULE SUMMARY" in line and "***" in line:
            module_summary_start = i
        elif "ENTRY LIST" in line and "***" in line:
            entry_list_start = i

    # ── Parse MODULE SUMMARY ────────────────────────────────────────────
    if module_summary_start is not None:
        # Find header line to determine column positions
        header_idx = None
        for i in range(module_summary_start, min(module_summary_start + 10, len(lines))):
            if "Module" in lines[i] and "ro code" in lines[i]:
                header_idx = i
                break

        if header_idx is not None:
            header_line = lines[header_idx]
            # Find column positions from header
            ro_code_col = header_line.find("ro code")
            ro_data_col = header_line.find("ro data")
            rw_data_col = header_line.find("rw data")

            current_module_group = ""
            i = header_idx + 2  # skip header + separator
            while i < len(lines):
                line = lines[i].rstrip("\r\n")

                # End of module summary section
                if "***" in line and line.strip().startswith("***"):
                    break

                stripped = line.strip()

                # Skip empty lines and separator lines
                if not stripped or stripped.startswith("---"):
                    i += 1
                    continue

                # Grand total line
                if "Grand Total:" in line:
                    parts = re.findall(r"[\d']+", line)
                    nums = [parse_iar_number(p) for p in parts]
                    if len(nums) >= 3:
                        result["grand_total"]["ro_code"] = nums[0]
                        result["grand_total"]["ro_data"] = nums[1]
                        result["grand_total"]["rw_data"] = nums[2]
                    elif len(nums) == 2:
                        result["grand_total"]["ro_code"] = nums[0]
                        result["grand_total"]["ro_data"] = nums[1]
                    elif len(nums) == 1:
                        result["grand_total"]["ro_code"] = nums[0]
                    i += 1
                    continue

                # Total line for a group — skip
                if stripped.startswith("Total:"):
                    i += 1
                    continue

                # Gaps / Linker created lines — skip
                if stripped.startswith("Gaps") or stripped.startswith("Linker created"):
                    i += 1
                    continue

                # Module group header: a line ending with : or : [N]
                # e.g., "C:\path\to\dir: [1]" or "dl7M_tlf.a: [5]" or "command line/config:"
                group_match = re.match(r"^(\S.*?):\s*(\[\d+\])?\s*$", line)
                if group_match and ".o" not in stripped.split()[0]:
                    current_module_group = group_match.group(1).strip()
                    # Simplify long paths to just the directory name
                    if "\\" in current_module_group or "/" in current_module_group:
                        parts_path = current_module_group.replace("\\", "/").split("/")
                        current_module_group = parts_path[-1] if parts_path[-1] else current_module_group
                    # Remove hash suffixes like _6603591812247902717.dir
                    current_module_group = re.sub(r"_\d{10,}\.dir", "", current_module_group)
                    i += 1
                    continue

                # Module data line: starts with whitespace, has a .o file,
                # then numbers aligned to the header columns
                # e.g., "    stm32f4xx_hal.o             144        8       12"
                # e.g., "    main.o                      380"
                mod_match = re.match(r"\s+(\S+\.o)\s+(.*)", line)
                if mod_match:
                    name = mod_match.group(1).replace(".o", "")
                    rest = mod_match.group(2).strip()
                    # Parse numbers from the rest of the line
                    nums = re.findall(r"[\d']+", rest)
                    nums = [parse_iar_number(n) for n in nums]

                    # Use column positions to determine which columns have data
                    # The numbers are right-aligned to their column headers
                    ro_code = 0
                    ro_data = 0
                    rw_data = 0

                    if ro_code_col >= 0 and ro_data_col >= 0 and rw_data_col >= 0:
                        # Extract text at each column region
                        # ro_code region: from ro_code_col to ro_data_col
                        # ro_data region: from ro_data_col to rw_data_col
                        # rw_data region: from rw_data_col to end
                        padded = line.ljust(rw_data_col + 10)
                        ro_code_text = padded[ro_code_col:ro_data_col].strip()
                        ro_data_text = padded[ro_data_col:rw_data_col].strip()
                        rw_data_text = padded[rw_data_col:].strip()

                        if ro_code_text:
                            ro_code = parse_iar_number(ro_code_text)
                        if ro_data_text:
                            ro_data = parse_iar_number(ro_data_text)
                        if rw_data_text:
                            rw_data = parse_iar_number(rw_data_text)
                    elif nums:
                        # Fallback: just use the numbers in order
                        ro_code = nums[0] if len(nums) > 0 else 0
                        ro_data = nums[1] if len(nums) > 1 else 0
                        rw_data = nums[2] if len(nums) > 2 else 0

                    result["modules"].append({
                        "name": name,
                        "group": current_module_group,
                        "ro_code": ro_code,
                        "ro_data": ro_data,
                        "rw_data": rw_data,
                        "total": ro_code + ro_data + rw_data,
                    })

                i += 1

    # ── Parse ENTRY LIST ────────────────────────────────────────────────
    if entry_list_start is not None:
        # Find header to get column positions
        header_idx = None
        for i in range(entry_list_start, min(entry_list_start + 10, len(lines))):
            if "Entry" in lines[i] and "Address" in lines[i] and "Size" in lines[i]:
                header_idx = i
                break

        if header_idx is not None:
            header_line = lines[header_idx]
            addr_col = header_line.find("Address")
            size_col = header_line.find("Size")
            type_col = header_line.find("Type")
            obj_col = header_line.find("Object")

            i = header_idx + 2  # skip header + separator
            while i < len(lines):
                line = lines[i].rstrip("\r\n")

                # End on footnotes section (lines like "[1] = ...")
                stripped = line.strip()
                if re.match(r"^\[\d+\]\s*=", stripped):
                    break
                # Skip empty lines
                if not stripped:
                    i += 1
                    continue

                # Try to parse as a full entry line
                # Entry lines have: Name  Address  Size  Type  Scope  Object
                # Some entries have no size (empty size column)
                entry_match = re.match(
                    r"(\S+)\s+(0x[\da-fA-F']+)\s+(0x[\da-fA-F]+)\s+(Code|Data)\s+(Gb|Lc|Wk)\s+(.*)",
                    stripped
                )
                if entry_match:
                    name = entry_match.group(1)
                    address = entry_match.group(2)
                    size_hex = entry_match.group(3)
                    entry_type = entry_match.group(4)
                    scope = entry_match.group(5)
                    obj = entry_match.group(6).strip()

                    size_bytes = int(size_hex, 16)
                    if size_bytes > 0:
                        result["entries"].append({
                            "name": name,
                            "address": address.replace("'", ""),
                            "size": size_bytes,
                            "type": entry_type,
                            "scope": scope,
                            "object": obj,
                        })
                    i += 1
                    continue

                # Check for a line with just a name (continuation/wrapped entry)
                # Next line will have the address/size/type data
                if stripped and not stripped.startswith("0x") and not stripped.startswith("-"):
                    pending_name = stripped
                    if i + 1 < len(lines):
                        next_line = lines[i + 1].strip()
                        cont_match = re.match(
                            r"(0x[\da-fA-F']+)\s+(0x[\da-fA-F]+)\s+(Code|Data)\s+(Gb|Lc|Wk)\s+(.*)",
                            next_line
                        )
                        if cont_match:
                            address = cont_match.group(1)
                            size_hex = cont_match.group(2)
                            entry_type = cont_match.group(3)
                            scope = cont_match.group(4)
                            obj = cont_match.group(5).strip()

                            size_bytes = int(size_hex, 16)
                            if size_bytes > 0:
                                result["entries"].append({
                                    "name": pending_name,
                                    "address": address.replace("'", ""),
                                    "size": size_bytes,
                                    "type": entry_type,
                                    "scope": scope,
                                    "object": obj,
                                })
                            i += 2  # skip both lines
                            continue

                i += 1

    # ── Parse final summary ─────────────────────────────────────────────
    for line in lines[-20:]:
        m = re.match(r"\s*([\d',]+)\s+bytes of readonly\s+code memory", line)
        if m:
            result["summary"]["readonly_code"] = parse_iar_number(m.group(1))
        m = re.match(r"\s*([\d',]+)\s+bytes of readonly\s+data memory", line)
        if m:
            result["summary"]["readonly_data"] = parse_iar_number(m.group(1))
        m = re.match(r"\s*([\d',]+)\s+bytes of readwrite data memory", line)
        if m:
            result["summary"]["readwrite_data"] = parse_iar_number(m.group(1))

    # Sort entries by size descending
    result["entries"].sort(key=lambda e: e["size"], reverse=True)

    return result


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────
def export_to_excel(data: dict, filepath: str, mcu_name: str = "STM32"):
    """Export parsed map data to a styled Excel workbook."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    wb = Workbook()

    # ── Styles ──────────────────────────────────────────────────────────
    header_font = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4A4E69", end_color="4A4E69", fill_type="solid")
    title_font = Font(name="Segoe UI", bold=True, size=14, color="22223B")
    subtitle_font = Font(name="Segoe UI", bold=False, size=11, color="4A4E69")
    data_font = Font(name="Segoe UI", size=10)
    number_font = Font(name="Consolas", size=10)
    border = Border(
        bottom=Side(style="thin", color="C9CCD5"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    accent_fill = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
    green_fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
    orange_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    light_row = PatternFill(start_color="F7F8FC", end_color="F7F8FC", fill_type="solid")

    total_flash = data["summary"]["readonly_code"] + data["summary"]["readonly_data"]
    total_ram = data["summary"]["readwrite_data"]

    # ── Sheet 1: Summary ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_properties.tabColor = "6C63FF"

    ws1.merge_cells("A1:D1")
    ws1["A1"] = "IAR EWARM Map File Analysis"
    ws1["A1"].font = title_font
    ws1["A1"].alignment = left_align

    ws1.merge_cells("A2:D2")
    ws1["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1["A2"].font = subtitle_font

    summary_rows = [
        ("Project", data["project_name"]),
        ("MCU / Platform", mcu_name),
        ("Toolchain", data["toolchain_info"]),
        ("", ""),
        ("MEMORY FOOTPRINT", ""),
        ("Read-only Code (Flash)", f"{data['summary']['readonly_code']:,} bytes"),
        ("Read-only Data (Flash)", f"{data['summary']['readonly_data']:,} bytes"),
        ("Total Flash", f"{total_flash:,} bytes"),
        ("Read-write Data (RAM)", f"{total_ram:,} bytes"),
        ("", ""),
        ("Total Modules", str(len(data["modules"]))),
        ("Total Functions/Entries", str(len(data["entries"]))),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, start=4):
        ws1.cell(row=row_idx, column=1, value=label).font = Font(name="Segoe UI", bold=True, size=11)
        ws1.cell(row=row_idx, column=2, value=value).font = Font(name="Segoe UI", size=11)
        if label in ("MEMORY FOOTPRINT",):
            ws1.cell(row=row_idx, column=1).font = Font(name="Segoe UI", bold=True, size=12, color="6C63FF")
        if label == "Total Flash":
            ws1.cell(row=row_idx, column=2).font = Font(name="Segoe UI", bold=True, size=11, color="27AE60")
        if label == "Read-write Data (RAM)":
            ws1.cell(row=row_idx, column=2).font = Font(name="Segoe UI", bold=True, size=11, color="F39C12")

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 45

    # ── Sheet 2: Module Breakdown ───────────────────────────────────────
    ws2 = wb.create_sheet("Module Breakdown")
    ws2.sheet_properties.tabColor = "2ECC71"

    headers2 = ["Module", "Group", "RO Code (bytes)", "RO Data (bytes)", "RW Data (bytes)", "Total (bytes)", "% of Flash"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    modules_sorted = sorted(data["modules"], key=lambda m: m["total"], reverse=True)
    for row_idx, mod in enumerate(modules_sorted, start=2):
        flash_pct = (mod["ro_code"] + mod["ro_data"]) / total_flash * 100 if total_flash > 0 else 0
        ws2.cell(row=row_idx, column=1, value=mod["name"]).font = data_font
        ws2.cell(row=row_idx, column=2, value=mod["group"]).font = data_font
        ws2.cell(row=row_idx, column=3, value=mod["ro_code"]).font = number_font
        ws2.cell(row=row_idx, column=4, value=mod["ro_data"]).font = number_font
        ws2.cell(row=row_idx, column=5, value=mod["rw_data"]).font = number_font
        ws2.cell(row=row_idx, column=6, value=mod["total"]).font = Font(name="Consolas", size=10, bold=True)
        ws2.cell(row=row_idx, column=7, value=round(flash_pct, 1)).font = number_font
        ws2.cell(row=row_idx, column=3).alignment = center_align
        ws2.cell(row=row_idx, column=4).alignment = center_align
        ws2.cell(row=row_idx, column=5).alignment = center_align
        ws2.cell(row=row_idx, column=6).alignment = center_align
        ws2.cell(row=row_idx, column=7).alignment = center_align
        if row_idx % 2 == 0:
            for c in range(1, 8):
                ws2.cell(row=row_idx, column=c).fill = light_row
        ws2.cell(row=row_idx, column=7).number_format = '0.0"%"'

    # Grand total row
    total_row = len(modules_sorted) + 2
    ws2.cell(row=total_row, column=1, value="GRAND TOTAL").font = Font(name="Segoe UI", bold=True, size=11)
    ws2.cell(row=total_row, column=3, value=data["grand_total"]["ro_code"]).font = Font(name="Consolas", bold=True, size=10)
    ws2.cell(row=total_row, column=4, value=data["grand_total"]["ro_data"]).font = Font(name="Consolas", bold=True, size=10)
    ws2.cell(row=total_row, column=5, value=data["grand_total"]["rw_data"]).font = Font(name="Consolas", bold=True, size=10)
    grand_total_sum = data["grand_total"]["ro_code"] + data["grand_total"]["ro_data"] + data["grand_total"]["rw_data"]
    ws2.cell(row=total_row, column=6, value=grand_total_sum).font = Font(name="Consolas", bold=True, size=10)

    for c in range(1, 8):
        ws2.cell(row=total_row, column=c).fill = PatternFill(start_color="E8E8F0", end_color="E8E8F0", fill_type="solid")

    col_widths2 = [30, 22, 16, 16, 16, 16, 12]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Function Breakdown ─────────────────────────────────────
    ws3 = wb.create_sheet("Function Breakdown")
    ws3.sheet_properties.tabColor = "F39C12"

    headers3 = ["Function / Entry", "Address", "Size (bytes)", "Type", "Scope", "Source Object"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row_idx, entry in enumerate(data["entries"], start=2):
        ws3.cell(row=row_idx, column=1, value=entry["name"]).font = data_font
        ws3.cell(row=row_idx, column=2, value=entry["address"]).font = Font(name="Consolas", size=10, color="6C63FF")
        ws3.cell(row=row_idx, column=3, value=entry["size"]).font = number_font
        ws3.cell(row=row_idx, column=4, value=entry["type"]).font = data_font
        ws3.cell(row=row_idx, column=5, value=entry["scope"]).font = data_font
        ws3.cell(row=row_idx, column=6, value=entry["object"]).font = data_font
        ws3.cell(row=row_idx, column=2).alignment = center_align
        ws3.cell(row=row_idx, column=3).alignment = center_align
        ws3.cell(row=row_idx, column=4).alignment = center_align
        ws3.cell(row=row_idx, column=5).alignment = center_align
        if row_idx % 2 == 0:
            for c in range(1, 7):
                ws3.cell(row=row_idx, column=c).fill = light_row

        # Highlight large functions (> 200 bytes)
        if entry["size"] >= 200:
            ws3.cell(row=row_idx, column=3).font = Font(name="Consolas", size=10, bold=True, color="E74C3C")

    col_widths3 = [35, 16, 14, 10, 10, 35]
    for i, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Auto-filter on all data sheets
    ws2.auto_filter.ref = f"A1:G{len(modules_sorted) + 1}"
    ws3.auto_filter.ref = f"A1:F{len(data['entries']) + 1}"

    wb.save(filepath)


# ─────────────────────────────────────────────────────────────────────────────
# GUI APPLICATION
# ─────────────────────────────────────────────────────────────────────────────
class MapAnalyzerApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("IAR Map Analyzer — STM32 Benchmarking Tool")
        self.root.geometry("1200x820")
        self.root.minsize(900, 650)
        self.root.configure(bg=COLORS["bg_dark"])

        self.data = None
        self.filepath = None

        self._apply_styles()
        self._build_ui()

    def _apply_styles(self):
        style = ttk.Style()
        style.theme_use("clam")

        # Treeview
        style.configure("Custom.Treeview",
                         background=COLORS["bg_card"],
                         foreground=COLORS["text_primary"],
                         fieldbackground=COLORS["bg_card"],
                         borderwidth=0,
                         font=(FONT_FAMILY, 10),
                         rowheight=28)
        style.configure("Custom.Treeview.Heading",
                         background=COLORS["bg_input"],
                         foreground=COLORS["text_primary"],
                         font=(FONT_FAMILY, 10, "bold"),
                         borderwidth=0,
                         relief="flat")
        style.map("Custom.Treeview",
                   background=[("selected", COLORS["accent"])],
                   foreground=[("selected", "#ffffff")])
        style.map("Custom.Treeview.Heading",
                   background=[("active", COLORS["border"])])

        # Notebook
        style.configure("Custom.TNotebook",
                         background=COLORS["bg_dark"],
                         borderwidth=0)
        style.configure("Custom.TNotebook.Tab",
                         background=COLORS["bg_card"],
                         foreground=COLORS["text_secondary"],
                         padding=[18, 8],
                         font=(FONT_FAMILY, 10, "bold"),
                         borderwidth=0)
        style.map("Custom.TNotebook.Tab",
                   background=[("selected", COLORS["accent"])],
                   foreground=[("selected", "#ffffff")])

        # Scrollbar
        style.configure("Custom.Vertical.TScrollbar",
                         background=COLORS["bg_input"],
                         troughcolor=COLORS["bg_card"],
                         borderwidth=0,
                         arrowsize=0)

    def _build_ui(self):
        # ── Top bar ─────────────────────────────────────────────────────
        top_frame = tk.Frame(self.root, bg=COLORS["bg_panel"], pady=12, padx=20)
        top_frame.pack(fill="x")

        title_label = tk.Label(top_frame,
                                text="⚡ IAR Map Analyzer",
                                bg=COLORS["bg_panel"],
                                fg=COLORS["accent"],
                                font=(FONT_FAMILY, 18, "bold"))
        title_label.pack(side="left")

        subtitle = tk.Label(top_frame,
                             text="STM32 Benchmarking Tool",
                             bg=COLORS["bg_panel"],
                             fg=COLORS["text_muted"],
                             font=(FONT_FAMILY, 11))
        subtitle.pack(side="left", padx=(12, 0), pady=(4, 0))

        # ── Controls bar ────────────────────────────────────────────────
        ctrl_frame = tk.Frame(self.root, bg=COLORS["bg_dark"], pady=10, padx=20)
        ctrl_frame.pack(fill="x")

        # Browse button
        self.browse_btn = tk.Button(ctrl_frame,
                                     text="📂  Browse Map File",
                                     bg=COLORS["accent"],
                                     fg="#ffffff",
                                     activebackground=COLORS["accent_hover"],
                                     activeforeground="#ffffff",
                                     font=(FONT_FAMILY, 11, "bold"),
                                     relief="flat",
                                     padx=20, pady=8,
                                     cursor="hand2",
                                     command=self._browse_file)
        self.browse_btn.pack(side="left")

        # MCU name entry
        tk.Label(ctrl_frame, text="MCU:", bg=COLORS["bg_dark"],
                 fg=COLORS["text_secondary"], font=(FONT_FAMILY, 10)).pack(side="left", padx=(20, 5))
        self.mcu_var = tk.StringVar(value="STM32")
        mcu_entry = tk.Entry(ctrl_frame, textvariable=self.mcu_var,
                              bg=COLORS["bg_input"], fg=COLORS["text_primary"],
                              insertbackground=COLORS["text_primary"],
                              font=(FONT_FAMILY, 10), relief="flat",
                              width=20, highlightthickness=1,
                              highlightbackground=COLORS["border"],
                              highlightcolor=COLORS["accent"])
        mcu_entry.pack(side="left")

        # File path label
        self.file_label = tk.Label(ctrl_frame, text="No file loaded",
                                    bg=COLORS["bg_dark"],
                                    fg=COLORS["text_muted"],
                                    font=(FONT_FAMILY, 9))
        self.file_label.pack(side="left", padx=(20, 0))

        # Export button
        self.export_btn = tk.Button(ctrl_frame,
                                     text="📊  Export to Excel",
                                     bg=COLORS["accent_green"],
                                     fg="#ffffff",
                                     activebackground="#3ddc84",
                                     activeforeground="#ffffff",
                                     font=(FONT_FAMILY, 11, "bold"),
                                     relief="flat",
                                     padx=20, pady=8,
                                     cursor="hand2",
                                     state="disabled",
                                     command=self._export_excel)
        self.export_btn.pack(side="right")

        # ── Summary cards ───────────────────────────────────────────────
        self.cards_frame = tk.Frame(self.root, bg=COLORS["bg_dark"], padx=20, pady=5)
        self.cards_frame.pack(fill="x")

        self.card_widgets = {}
        card_defs = [
            ("ro_code",  "Read-Only Code",     "0",  COLORS["accent"]),
            ("ro_data",  "Read-Only Data",     "0",  COLORS["accent_orange"]),
            ("rw_data",  "Read-Write Data",    "0",  COLORS["accent_green"]),
            ("flash",    "Total Flash",        "0",  "#e74c3c"),
            ("ram",      "Total RAM",          "0",  "#9b59b6"),
        ]
        for key, label, default, color in card_defs:
            card = self._create_card(self.cards_frame, label, default, color)
            card.pack(side="left", expand=True, fill="x", padx=4)
            self.card_widgets[key] = card

        # ── Notebook with tables ────────────────────────────────────────
        nb_frame = tk.Frame(self.root, bg=COLORS["bg_dark"], padx=20, pady=8)
        nb_frame.pack(fill="both", expand=True)

        self.notebook = ttk.Notebook(nb_frame, style="Custom.TNotebook")
        self.notebook.pack(fill="both", expand=True)

        # Tab 1: Module Breakdown
        mod_frame = tk.Frame(self.notebook, bg=COLORS["bg_card"])
        self.notebook.add(mod_frame, text="  📦 Module Breakdown  ")

        mod_cols = ("group", "ro_code", "ro_data", "rw_data", "total", "pct")
        self.mod_tree = ttk.Treeview(mod_frame, columns=mod_cols,
                                      show="tree headings", style="Custom.Treeview")
        self.mod_tree.heading("#0", text="Module", anchor="w")
        self.mod_tree.heading("group", text="Group", anchor="w")
        self.mod_tree.heading("ro_code", text="RO Code", anchor="center")
        self.mod_tree.heading("ro_data", text="RO Data", anchor="center")
        self.mod_tree.heading("rw_data", text="RW Data", anchor="center")
        self.mod_tree.heading("total", text="Total", anchor="center")
        self.mod_tree.heading("pct", text="% Flash", anchor="center")

        self.mod_tree.column("#0", width=200, minwidth=120)
        self.mod_tree.column("group", width=160, minwidth=80)
        self.mod_tree.column("ro_code", width=100, anchor="center")
        self.mod_tree.column("ro_data", width=100, anchor="center")
        self.mod_tree.column("rw_data", width=100, anchor="center")
        self.mod_tree.column("total", width=100, anchor="center")
        self.mod_tree.column("pct", width=80, anchor="center")

        mod_scroll = ttk.Scrollbar(mod_frame, orient="vertical",
                                    command=self.mod_tree.yview,
                                    style="Custom.Vertical.TScrollbar")
        self.mod_tree.configure(yscrollcommand=mod_scroll.set)
        self.mod_tree.pack(side="left", fill="both", expand=True)
        mod_scroll.pack(side="right", fill="y")

        # Tab 2: Function Breakdown
        func_frame = tk.Frame(self.notebook, bg=COLORS["bg_card"])
        self.notebook.add(func_frame, text="  🔧 Function Breakdown  ")

        func_cols = ("address", "size", "type", "scope", "object")
        self.func_tree = ttk.Treeview(func_frame, columns=func_cols,
                                       show="tree headings", style="Custom.Treeview")
        self.func_tree.heading("#0", text="Function / Entry", anchor="w")
        self.func_tree.heading("address", text="Address", anchor="center")
        self.func_tree.heading("size", text="Size (bytes)", anchor="center")
        self.func_tree.heading("type", text="Type", anchor="center")
        self.func_tree.heading("scope", text="Scope", anchor="center")
        self.func_tree.heading("object", text="Source", anchor="w")

        self.func_tree.column("#0", width=280, minwidth=150)
        self.func_tree.column("address", width=130, anchor="center")
        self.func_tree.column("size", width=110, anchor="center")
        self.func_tree.column("type", width=70, anchor="center")
        self.func_tree.column("scope", width=60, anchor="center")
        self.func_tree.column("object", width=250)

        func_scroll = ttk.Scrollbar(func_frame, orient="vertical",
                                     command=self.func_tree.yview,
                                     style="Custom.Vertical.TScrollbar")
        self.func_tree.configure(yscrollcommand=func_scroll.set)
        self.func_tree.pack(side="left", fill="both", expand=True)
        func_scroll.pack(side="right", fill="y")

        # ── Status bar ──────────────────────────────────────────────────
        self.status_bar = tk.Label(self.root,
                                    text="Ready — Load a .map file to begin analysis",
                                    bg=COLORS["bg_panel"],
                                    fg=COLORS["text_muted"],
                                    font=(FONT_FAMILY, 9),
                                    anchor="w", padx=20, pady=6)
        self.status_bar.pack(fill="x", side="bottom")

    def _create_card(self, parent, label, value, accent_color):
        card = tk.Frame(parent, bg=COLORS["bg_card"], padx=16, pady=12,
                         highlightbackground=COLORS["border"],
                         highlightthickness=1)

        # Top accent line
        accent_line = tk.Frame(card, bg=accent_color, height=3)
        accent_line.pack(fill="x")

        tk.Label(card, text=label,
                 bg=COLORS["bg_card"], fg=COLORS["text_secondary"],
                 font=(FONT_FAMILY, 9)).pack(anchor="w", pady=(6, 0))

        val_label = tk.Label(card, text=value,
                              bg=COLORS["bg_card"], fg=COLORS["text_primary"],
                              font=(FONT_FAMILY, 16, "bold"))
        val_label.pack(anchor="w")
        card._val_label = val_label

        unit_label = tk.Label(card, text="bytes",
                               bg=COLORS["bg_card"], fg=COLORS["text_muted"],
                               font=(FONT_FAMILY, 8))
        unit_label.pack(anchor="w")

        return card

    def _update_card(self, key, value):
        card = self.card_widgets[key]
        card._val_label.config(text=f"{value:,}")

    def _browse_file(self):
        filepath = filedialog.askopenfilename(
            title="Select IAR EWARM Map File",
            filetypes=[("Map files", "*.map"), ("All files", "*.*")]
        )
        if not filepath:
            return

        self.filepath = filepath
        self.file_label.config(text=os.path.basename(filepath), fg=COLORS["text_primary"])
        self.status_bar.config(text=f"Parsing: {filepath}")
        self.root.update_idletasks()

        try:
            self.data = parse_map_file(filepath)
            self._populate_ui()
            self.export_btn.config(state="normal")
            self.status_bar.config(
                text=f"✅ Loaded: {self.data['project_name']} — "
                     f"{len(self.data['modules'])} modules, "
                     f"{len(self.data['entries'])} entries",
                fg=COLORS["accent_green"]
            )
        except Exception as e:
            messagebox.showerror("Parse Error", f"Failed to parse map file:\n{e}")
            self.status_bar.config(text=f"❌ Error: {e}", fg=COLORS["accent_red"])

    def _populate_ui(self):
        d = self.data

        # Update cards
        total_flash = d["summary"]["readonly_code"] + d["summary"]["readonly_data"]
        total_ram = d["summary"]["readwrite_data"]
        self._update_card("ro_code", d["summary"]["readonly_code"])
        self._update_card("ro_data", d["summary"]["readonly_data"])
        self._update_card("rw_data", d["summary"]["readwrite_data"])
        self._update_card("flash", total_flash)
        self._update_card("ram", total_ram)

        # Populate module tree
        self.mod_tree.delete(*self.mod_tree.get_children())
        modules_sorted = sorted(d["modules"], key=lambda m: m["total"], reverse=True)
        for mod in modules_sorted:
            flash_pct = (mod["ro_code"] + mod["ro_data"]) / total_flash * 100 if total_flash > 0 else 0
            self.mod_tree.insert("", "end", text=mod["name"],
                                  values=(mod["group"],
                                          f'{mod["ro_code"]:,}',
                                          f'{mod["ro_data"]:,}',
                                          f'{mod["rw_data"]:,}',
                                          f'{mod["total"]:,}',
                                          f'{flash_pct:.1f}%'))

        # Populate function tree
        self.func_tree.delete(*self.func_tree.get_children())
        for entry in d["entries"]:
            self.func_tree.insert("", "end", text=entry["name"],
                                   values=(entry["address"],
                                           f'{entry["size"]:,}',
                                           entry["type"],
                                           entry["scope"],
                                           entry["object"]))

    def _export_excel(self):
        if not self.data:
            messagebox.showwarning("No Data", "Please load a map file first.")
            return

        if not HAS_OPENPYXL:
            messagebox.showerror(
                "Missing Dependency",
                "openpyxl is required for Excel export.\n\nInstall with:\n  pip install openpyxl"
            )
            return

        default_name = f"{self.data['project_name']}_footprint_{self.mcu_var.get()}.xlsx"
        filepath = filedialog.asksaveasfilename(
            title="Export to Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not filepath:
            return

        try:
            export_to_excel(self.data, filepath, self.mcu_var.get())
            self.status_bar.config(
                text=f"📊 Exported to: {os.path.basename(filepath)}",
                fg=COLORS["accent_green"]
            )
            messagebox.showinfo("Export Complete",
                                f"Excel file saved to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export:\n{e}")
            self.status_bar.config(text=f"❌ Export error: {e}", fg=COLORS["accent_red"])


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    app = MapAnalyzerApp(root)
    root.mainloop()
